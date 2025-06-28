from flask import Blueprint, request, jsonify
from app.models.cows import Cow
from app.database.database import db
from fpdf import FPDF
from app.models.daily_milk_summary import DailyMilkSummary  # Add this line
from flask import send_file
from io import BytesIO
import pandas as pd

cow_bp = Blueprint('cow', __name__)

@cow_bp.route('/add', methods=['POST'])
def add_cow():
    """
    Menambahkan sapi baru ke database.
    """
    try:
        data = request.get_json()
        name = data.get('name')
        birth = data.get('birth')
        breed = data.get('breed')
        lactation_phase = data.get('lactation_phase')
        weight = data.get('weight')
        gender = data.get('gender')

        # Validasi input
        if not name or not birth or not breed or not gender:
            return jsonify({"error": "Missing required fields"}), 400

        # Buat instance Cow baru
        new_cow = Cow(
            name=name,
            birth=birth,
            breed=breed,
            lactation_phase=lactation_phase,
            weight=weight,
            gender=gender
        )

        # Simpan ke database
        db.session.add(new_cow)
        db.session.commit()

        return jsonify({"message": "Cow added successfully", "cow": {
            "id": new_cow.id,
            "name": new_cow.name,
            "birth": new_cow.birth,
            "breed": new_cow.breed,
            "lactation_phase": new_cow.lactation_phase,
            "weight": new_cow.weight,
            "gender": new_cow.gender
        }}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@cow_bp.route('/<int:cow_id>', methods=['GET'])
def get_cow_by_id(cow_id):
    """
    Mendapatkan data sapi berdasarkan ID.
    """
    try:
        cow = Cow.query.get(cow_id)
        if not cow:
            return jsonify({"error": "Cow not found"}), 404

        return jsonify({"cow": {
            "id": cow.id,
            "name": cow.name,
            "birth": cow.birth,
            "breed": cow.breed,
            "lactation_phase": cow.lactation_phase,
            "weight": cow.weight,
            "gender": cow.gender
        }}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@cow_bp.route('/list', methods=['GET'])
def list_cows():
    """
    Mendapatkan daftar semua sapi.
    """
    try:
        cows = Cow.query.all()
        cows_list = [{
            "id": cow.id,
            "name": cow.name,
            "birth": cow.birth,
            "breed": cow.breed,
            "lactation_phase": cow.lactation_phase,
            "weight": cow.weight,
            "gender": cow.gender
        } for cow in cows]

        return jsonify({"cows": cows_list}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@cow_bp.route('/update/<int:cow_id>', methods=['PUT'])
def update_cow(cow_id):
    """
    Memperbarui data sapi berdasarkan ID.
    """
    try:
        data = request.get_json()
        cow = Cow.query.get(cow_id)
        if not cow:
            return jsonify({"error": "Cow not found"}), 404

        # Perbarui data sapi
        cow.name = data.get('name', cow.name)
        cow.birth = data.get('birth', cow.birth)
        cow.breed = data.get('breed', cow.breed)
        cow.lactation_phase = data.get('lactation_phase', cow.lactation_phase)
        cow.weight = data.get('weight', cow.weight)
        cow.gender = data.get('gender', cow.gender)

        db.session.commit()

        return jsonify({"message": "Cow updated successfully", "cow": {
            "id": cow.id,
            "name": cow.name,
            "birth": cow.birth,
            "breed": cow.breed,
            "lactation_phase": cow.lactation_phase,
            "weight": cow.weight,
            "gender": cow.gender
        }}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@cow_bp.route('/delete/<int:cow_id>', methods=['DELETE'])
def delete_cow(cow_id):
    """
    Menghapus data sapi berdasarkan ID, sekaligus menghapus data produksi susu & milking_sessions terkait (paksa hapus).
    """
    try:
        print("="*50)
        print(f"[DEBUG] [DELETE COW] Mulai proses hapus sapi dengan ID: {cow_id}")

        cow = Cow.query.get(cow_id)
        if not cow:
            print(f"[DEBUG] [DELETE COW] Sapi dengan ID {cow_id} TIDAK DITEMUKAN di database.")
            print("="*50)
            return jsonify({"error": "Cow not found"}), 404

        print(f"[DEBUG] [DELETE COW] Sapi ditemukan: {cow.name} (ID: {cow.id})")

        # Import text for SQL statements
        from sqlalchemy import text

        # Disable foreign key checks temporarily
        db.session.execute(text("SET FOREIGN_KEY_CHECKS = 0"))
        
        # Hapus semua data milking_sessions terkait
        from app.models.milking_sessions import MilkingSession
        milking_sessions = MilkingSession.query.filter_by(cow_id=cow_id).all()
        print(f"[DEBUG] [DELETE COW] Jumlah milking_sessions terkait sapi ID {cow_id}: {len(milking_sessions)}")
        for session in milking_sessions:
            db.session.delete(session)
        if milking_sessions:
            print(f"[DEBUG] [DELETE COW] Semua milking_sessions terkait telah dihapus.")

        # Hapus semua data produksi susu terkait
        related_records = DailyMilkSummary.query.filter_by(cow_id=cow_id).all()
        print(f"[DEBUG] [DELETE COW] Jumlah data produksi susu terkait sapi ID {cow_id}: {len(related_records)}")
        for record in related_records:
            db.session.delete(record)
        if related_records:
            print(f"[DEBUG] [DELETE COW] Semua data produksi susu terkait telah dihapus.")

        # Force delete related feed schedules using raw SQL
        db.session.execute(text("DELETE FROM daily_feed_schedule WHERE cow_id = :cow_id"), {"cow_id": cow_id})
        print(f"[DEBUG] [DELETE COW] Semua daily_feed_schedule terkait telah dihapus dengan paksa.")

        # Hapus data sapi
        db.session.delete(cow)
        
        # Re-enable foreign key checks
        db.session.execute(text("SET FOREIGN_KEY_CHECKS = 1"))
        
        db.session.commit()
        print(f"[DEBUG] [DELETE COW] Sapi ID {cow_id} beserta data terkait berhasil dihapus dari database.")
        print("="*50)

        return jsonify({"message": "Cow and all related records deleted successfully"}), 200

    except Exception as e:
        db.session.rollback()
        # Re-enable foreign key checks in case of error
        try:
            db.session.execute(text("SET FOREIGN_KEY_CHECKS = 1"))
            db.session.commit()
        except:
            pass
        print(f"[ERROR] [DELETE COW] Terjadi error saat menghapus sapi ID {cow_id}: {e}")
        print("="*50)
        return jsonify({"error": str(e)}), 500

@cow_bp.route('/export/pdf', methods=['GET'])
def export_cows_pdf():
    """
    Mengekspor data sapi ke dalam file PDF.
    """
    try:
        # Ambil semua data sapi dari database
        cows = Cow.query.all()

        # Buat PDF
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Tambahkan deskripsi di bagian atas
        pdf.set_font("Arial", style="B", size=16)
        pdf.cell(200, 10, txt="Laporan Data Sapi", ln=True, align='C')
        pdf.ln(5)
        pdf.set_font("Arial", size=10)
        pdf.cell(200, 10, txt="Berikut adalah daftar sapi yang terdaftar dalam sistem.", ln=True, align='C')
        pdf.ln(10)

        # Tambahkan header tabel dengan warna latar belakang
        pdf.set_fill_color(173, 216, 230)  # Warna biru muda (RGB)
        pdf.set_text_color(0, 0, 0)  # Warna teks hitam
        pdf.set_font("Arial", style="B", size=10)
        pdf.cell(20, 10, "NO", border=1, align='C', fill=True)
        pdf.cell(40, 10, "Name", border=1, align='C', fill=True)
        pdf.cell(40, 10, "Breed", border=1, align='C', fill=True)
        pdf.cell(40, 10, "Gender", border=1, align='C', fill=True)
        pdf.cell(50, 10, "Lactation Phase", border=1, align='C', fill=True)
        pdf.ln()

        # Isi data
        pdf.set_font("Arial", size=10)
        for idx, cow in enumerate(cows, start=1):
            pdf.cell(20, 10, str(idx), border=1, align='C')
            pdf.cell(40, 10, cow.name, border=1)
            pdf.cell(40, 10, cow.breed, border=1)
            pdf.cell(40, 10, cow.gender, border=1)
            pdf.cell(50, 10, cow.lactation_phase or "-", border=1)
            pdf.ln()

       # Simpan PDF ke buffer
        buffer = BytesIO()
        pdf_bytes = pdf.output(dest='S').encode('latin1')
        buffer.write(pdf_bytes)
        buffer.seek(0)

        return send_file(buffer, as_attachment=True, download_name="cows.pdf", mimetype='application/pdf')

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@cow_bp.route('/export/excel', methods=['GET'])
def export_cows_excel():
    """
    Mengekspor data sapi ke dalam file Excel.
    """
    try:
        # Ambil semua data sapi dari database
        cows = Cow.query.all()

        # Format data sapi menjadi list of dictionaries
        cows_list = [{
            "NO": idx,
            "Name": cow.name,
            "Breed": cow.breed,
            "Gender": cow.gender,
            "Lactation Phase": cow.lactation_phase or "-",
            "Weight": cow.weight or "-",
            "Birth": cow.birth
        } for idx, cow in enumerate(cows, start=1)]

        # Buat DataFrame
        df = pd.DataFrame(cows_list)

        # Simpan ke buffer dengan pewarnaan
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Cows')

            # Akses workbook dan worksheet untuk styling
            workbook = writer.book
            worksheet = writer.sheets['Cows']

            # Tambahkan pewarnaan header
            from openpyxl.styles import Font, PatternFill
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            header_font = Font(bold=True)

            for cell in worksheet[1]:  # Baris pertama adalah header
                cell.fill = header_fill
                cell.font = header_font

            # Atur lebar kolom secara otomatis
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        buffer.seek(0)

        return send_file(buffer, as_attachment=True, download_name="cows.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({"error": str(e)}), 500