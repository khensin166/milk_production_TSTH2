from flask import Blueprint, request, jsonify
from werkzeug.security import generate_password_hash
from app.models.users import User
from app.models.roles import Role
from app.database.database import db
from fpdf import FPDF
from flask import send_file
from io import BytesIO
import pandas as pd
from werkzeug.security import check_password_hash
import logging
import traceback

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

user_bp = Blueprint('user', __name__)

@user_bp.route('/list', methods=['GET'])
def get_all_users():
    try:
        # Ambil semua data pengguna dari database
        users = User.query.all()

        # Format data pengguna menjadi list of dictionaries
        users_list = [{
            "id": user.id,
            "name": user.name,  # Tambahkan name
            "username": user.username,
            "email": user.email,
            "contact": user.contact,
            "religion": user.religion,
            "role_id": user.role_id,
            "birth": user.birth,  # Tambahkan birth
            "token": user.token
        } for user in users]

        return jsonify({"users": users_list}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@user_bp.route('/<int:user_id>', methods=['GET'])
def get_user_by_id(user_id):
    try:
        # Cari user berdasarkan user_id
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Format data pengguna menjadi dictionary
        user_data = {
            "id": user.id,
            "name": user.name,  # Tambahkan name
            "username": user.username,
            "email": user.email,
            "contact": user.contact,
            "religion": user.religion,
            "role_id": user.role_id,
            "birth": user.birth,  # Tambahkan birth
            "token": user.token,
        }

        return jsonify({"user": user_data}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@user_bp.route('/add', methods=['POST'])
def add_user():
    try:
        # Ambil data dari request body
        data = request.get_json()
        name = data.get('name')  # Tambahkan name
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        contact = data.get('contact')
        religion = data.get('religion')
        birth = data.get('birth')  # Tambahkan birth
        role_id = data.get('role_id')  # Pastikan role_id sesuai dengan user
        token = data.get('token', None)

        # Validasi data
        if not name or not username or not email or not password or not role_id:
            return jsonify({"error": "Missing required fields"}), 400

        # Cek apakah username atau email sudah ada
        if User.query.filter_by(username=username).first():
            return jsonify({"error": "Username already exists"}), 400
        if User.query.filter_by(email=email).first():
            return jsonify({"error": "Email already exists"}), 400

        # Validasi role_id
        role = Role.query.get(role_id)
        if not role:
            return jsonify({"error": "Invalid role_id"}), 400

        # Hash password
        hashed_password = generate_password_hash(password)

        # Buat instance User baru
        new_user = User(
            name=name,  # Tambahkan name
            username=username,
            email=email,
            password=hashed_password,
            contact=contact,
            religion=religion,
            birth=birth,  # Tambahkan birth
            role_id=role_id,
            token=token
        )

        # Simpan ke database
        db.session.add(new_user)
        db.session.commit()

        return jsonify({"message": "User added successfully", "user": {
            "name": new_user.name,  # Tambahkan name
            "username": new_user.username,
            "email": new_user.email,
            "contact": new_user.contact,
            "religion": new_user.religion,
            "birth": new_user.birth,  # Tambahkan birth
            "role_id": new_user.role_id
        }}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
        
@user_bp.route('/export/pdf', methods=['GET'])
def export_users_pdf():
    try:
        # Ambil semua data pengguna dari database
        users = User.query.join(Role, User.role_id == Role.id).add_columns(
            User.id, User.name, User.username, User.email, User.contact,
            User.religion, User.birth, Role.name.label('role_name')
        ).all()

        # Buat PDF
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Tambahkan deskripsi di bagian atas
        pdf.set_font("Arial", style="B", size=16)
        pdf.cell(200, 10, txt="Laporan Data Pengguna", ln=True, align='C')
        pdf.ln(5)
        pdf.set_font("Arial", size=10)
        pdf.cell(200, 10, txt="Berikut adalah daftar pengguna yang terdaftar dalam sistem.", ln=True, align='C')
        pdf.ln(10)

        # Tambahkan header tabel dengan warna latar belakang
        pdf.set_fill_color(173, 216, 230)  # Warna biru muda (RGB)
        pdf.set_text_color(0, 0, 0)  # Warna teks hitam
        pdf.set_font("Arial", style="B", size=10)
        pdf.cell(20, 10, "NO", border=1, align='C', fill=True)  # Ganti ID menjadi NO
        pdf.cell(40, 10, "Name", border=1, align='C', fill=True)
        pdf.cell(40, 10, "Username", border=1, align='C', fill=True)
        pdf.cell(50, 10, "Email", border=1, align='C', fill=True)
        pdf.cell(40, 10, "Role", border=1, align='C', fill=True)
        pdf.ln()

        # Isi data
        pdf.set_font("Arial", size=10)
        for idx, user in enumerate(users, start=1):  # Gunakan enumerate untuk nomor urut
            pdf.cell(20, 10, str(idx), border=1, align='C')  # Tampilkan nomor urut
            pdf.cell(40, 10, user.name, border=1)
            pdf.cell(40, 10, user.username, border=1)
            pdf.cell(50, 10, user.email, border=1)
            pdf.cell(40, 10, user.role_name, border=1)
            pdf.ln()

        # Simpan PDF ke buffer (di luar loop)
        buffer = BytesIO()
        pdf_bytes = pdf.output(dest='S').encode('latin1')
        buffer.write(pdf_bytes)
        buffer.seek(0)

        return send_file(buffer, as_attachment=True, download_name="users.pdf", mimetype='application/pdf')

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@user_bp.route('/export/excel', methods=['GET'])
def export_users_excel():
    try:
        # Ambil semua data pengguna dari database dengan join ke tabel Role
        users = User.query.join(Role, User.role_id == Role.id).add_columns(
            User.id, User.name, User.username, User.email, User.contact,
            User.religion, User.birth, Role.name.label('role_name')
        ).all()

        # Format data pengguna menjadi list of dictionaries
        users_list = [{
            "NO": idx,  # Tambahkan nomor urut
            "Name": user.name,
            "Username": user.username,
            "Email": user.email,
            "Contact": user.contact,
            "Religion": user.religion,
            "Role": user.role_name,  # Ganti Role ID dengan nama role
            "Birth": user.birth
        } for idx, user in enumerate(users, start=1)]  # Gunakan enumerate untuk nomor urut

        # Buat DataFrame
        df = pd.DataFrame(users_list)

        # Simpan ke buffer dengan pewarnaan
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Users')

            # Akses workbook dan worksheet untuk styling
            workbook = writer.book
            worksheet = writer.sheets['Users']

            # Tambahkan pewarnaan header
            from openpyxl.styles import Font, PatternFill
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            header_font = Font(bold=True)

            for cell in worksheet[1]:  # Baris pertama adalah header
                cell.fill = header_fill
                cell.font = header_font

            # Atur lebar kolom secara otomatis
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        buffer.seek(0)

        return send_file(buffer, as_attachment=True, download_name="users.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    

@user_bp.route('/delete/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):    
    try:
        logger.info(f"Starting deletion process for user ID: {user_id}")
        
        # Find user by ID
        user = User.query.get(user_id)
        if not user:
            logger.warning(f"User ID {user_id} not found")
            return jsonify({"error": "User not found"}), 404

        logger.info(f"Found user: {user.id} ({user.username})")
        
        # Begin transaction
        deletion_summary = {}   
        
        try:
            # Check database for foreign keys to users table
            from sqlalchemy import text, inspect
            inspector = inspect(db.engine)
            
            # Get all tables in the database
            all_tables = inspector.get_table_names()
            
            # List tables that need special handling
            special_handling_tables = [
                "feed_stock", "product_stock", "daily_feed_schedule", "feed", 
                "feed_type", "nutritions", "feed_stock_history", "product_type",
                "order_item", "daily_feed_items"
            ]
            
            # Handle potentially nullable columns first (especially those containing "_by")
            logger.info("Step 1: Handling potentially nullable columns")
            for table_name in all_tables:
                try:
                    # Find all foreign keys that reference the users table
                    fk_columns = []
                    for fk in inspector.get_foreign_keys(table_name):
                        if fk.get('referred_table') == 'users':
                            fk_columns.extend(fk.get('constrained_columns', []))
                    
                    # First try to handle columns that contain "_by" (like "created_by", "updated_by")
                    # These are usually nullable
                    for column in fk_columns:
                        if "_by" in column:
                            logger.info(f"Processing {table_name}.{column} references to user {user_id}")
                            try:
                                query = text(f"UPDATE {table_name} SET {column} = NULL WHERE {column} = :user_id")
                                result = db.session.execute(query, {"user_id": user_id})
                                deletion_summary[f"{table_name}.{column}_nullified"] = result.rowcount
                                logger.info(f"Nullified {result.rowcount} records in {table_name}.{column}")
                                db.session.commit()
                            except Exception as e:
                                logger.warning(f"Could not nullify {table_name}.{column}: {str(e)}")
                                db.session.rollback()
                except Exception as e:
                    logger.warning(f"Error processing columns for {table_name}: {str(e)}")
            
            # Handle special tables that need specific processing order
            logger.info("Step 2: Handling special tables")
            
            # Process feed_stock_history
            try:
                if "feed_stock_history" in all_tables:
                    logger.info(f"Processing feed_stock_history references to user {user_id}")
                    
                    # Check for any not-yet-nullified columns
                    fk_columns = []
                    for fk in inspector.get_foreign_keys("feed_stock_history"):
                        if fk.get('referred_table') == 'users':
                            fk_columns.extend(fk.get('constrained_columns', []))
                    
                    for column in fk_columns:
                        # Try to nullify first
                        try:
                            query = text(f"UPDATE feed_stock_history SET {column} = NULL WHERE {column} = :user_id")
                            result = db.session.execute(query, {"user_id": user_id})
                            deletion_summary[f"feed_stock_history.{column}_nullified"] = result.rowcount
                            logger.info(f"Nullified {result.rowcount} records in feed_stock_history.{column}")
                            db.session.commit()
                        except Exception as e:
                            logger.warning(f"Could not nullify feed_stock_history.{column}: {str(e)}")
                            db.session.rollback()
                            
                            # If nullification fails, delete
                            query = text(f"DELETE FROM feed_stock_history WHERE {column} = :user_id")
                            result = db.session.execute(query, {"user_id": user_id})
                            deletion_summary[f"feed_stock_history.{column}_deleted"] = result.rowcount
                            logger.info(f"Deleted {result.rowcount} records from feed_stock_history where {column} = {user_id}")
                            db.session.commit()
            except Exception as e:
                logger.error(f"Error handling feed_stock_history records: {str(e)}")
                db.session.rollback()
                return jsonify({
                    "status": "error",
                    "message": "Failed to handle feed_stock_history records",
                    "details": str(e)
                }), 500
            
            # Process nutritions table
            try:
                if "nutritions" in all_tables:
                    logger.info(f"Processing nutritions references to user {user_id}")
                    
                    # Check for any not-yet-nullified columns
                    fk_columns = []
                    for fk in inspector.get_foreign_keys("nutritions"):
                        if fk.get('referred_table') == 'users':
                            fk_columns.extend(fk.get('constrained_columns', []))
                    
                    for column in fk_columns:
                        if "_by" not in column:  # Skip columns already handled
                            # Try to nullify first
                            try:
                                query = text(f"UPDATE nutritions SET {column} = NULL WHERE {column} = :user_id")
                                result = db.session.execute(query, {"user_id": user_id})
                                deletion_summary[f"nutritions.{column}_nullified"] = result.rowcount
                                logger.info(f"Nullified {result.rowcount} records in nutritions.{column}")
                                db.session.commit()
                            except Exception as e:
                                logger.warning(f"Could not nullify nutritions.{column}: {str(e)}")
                                db.session.rollback()
                                
                                # If nullification fails, delete
                                query = text(f"DELETE FROM nutritions WHERE {column} = :user_id")
                                result = db.session.execute(query, {"user_id": user_id})
                                deletion_summary[f"nutritions.{column}_deleted"] = result.rowcount
                                logger.info(f"Deleted {result.rowcount} records from nutritions where {column} = {user_id}")
                                db.session.commit()
            except Exception as e:
                logger.error(f"Error handling nutritions records: {str(e)}")
                db.session.rollback()
                return jsonify({
                    "status": "error",
                    "message": "Failed to handle nutritions records",
                    "details": str(e)
                }), 500
            
            # Process feed_type table 
            try:
                if "feed_type" in all_tables:
                    logger.info(f"Processing feed_type references to user {user_id}")
                    
                    # Check for any not-yet-nullified columns
                    fk_columns = []
                    for fk in inspector.get_foreign_keys("feed_type"):
                        if fk.get('referred_table') == 'users':
                            fk_columns.extend(fk.get('constrained_columns', []))
                    
                    for column in fk_columns:
                        if "_by" not in column:  # Skip columns already handled
                            # Try to nullify first
                            try:
                                query = text(f"UPDATE feed_type SET {column} = NULL WHERE {column} = :user_id")
                                result = db.session.execute(query, {"user_id": user_id})
                                deletion_summary[f"feed_type.{column}_nullified"] = result.rowcount
                                logger.info(f"Nullified {result.rowcount} records in feed_type.{column}")
                                db.session.commit()
                            except Exception as e:
                                logger.warning(f"Could not nullify feed_type.{column}: {str(e)}")
                                db.session.rollback()
                                
                                # If nullification fails, delete
                                query = text(f"DELETE FROM feed_type WHERE {column} = :user_id")
                                result = db.session.execute(query, {"user_id": user_id})
                                deletion_summary[f"feed_type.{column}_deleted"] = result.rowcount
                                logger.info(f"Deleted {result.rowcount} records from feed_type where {column} = {user_id}")
                                db.session.commit()
            except Exception as e:
                logger.error(f"Error handling feed_type records: {str(e)}")
                db.session.rollback()
                return jsonify({
                    "status": "error",
                    "message": "Failed to handle feed_type records",
                    "details": str(e)
                }), 500
            
            # Process feed table
            try:
                if "feed" in all_tables:
                    logger.info(f"Processing feed references to user {user_id}")
                    
                    # Check for any not-yet-nullified columns
                    fk_columns = []
                    for fk in inspector.get_foreign_keys("feed"):
                        if fk.get('referred_table') == 'users':
                            fk_columns.extend(fk.get('constrained_columns', []))
                    
                    for column in fk_columns:
                        if "_by" not in column:  # Skip columns already handled
                            # Try to nullify first
                            try:
                                query = text(f"UPDATE feed SET {column} = NULL WHERE {column} = :user_id")
                                result = db.session.execute(query, {"user_id": user_id})
                                deletion_summary[f"feed.{column}_nullified"] = result.rowcount
                                logger.info(f"Nullified {result.rowcount} records in feed.{column}")
                                db.session.commit()
                            except Exception as e:
                                logger.warning(f"Could not nullify feed.{column}: {str(e)}")
                                db.session.rollback()
                                
                                # If nullification fails, delete
                                query = text(f"DELETE FROM feed WHERE {column} = :user_id")
                                result = db.session.execute(query, {"user_id": user_id})
                                deletion_summary[f"feed.{column}_deleted"] = result.rowcount
                                logger.info(f"Deleted {result.rowcount} records from feed where {column} = {user_id}")
                                db.session.commit()
            except Exception as e:
                logger.error(f"Error handling feed records: {str(e)}")
                db.session.rollback()
                return jsonify({
                    "status": "error",
                    "message": "Failed to handle feed records",
                    "details": str(e)
                }), 500
            
            # Process daily_feed_schedule table
            try:
                if "daily_feed_schedule" in all_tables:
                    logger.info(f"Processing daily_feed_schedule references to user {user_id}")
                    
                    # Check for any not-yet-nullified columns
                    fk_columns = []
                    for fk in inspector.get_foreign_keys("daily_feed_schedule"):
                        if fk.get('referred_table') == 'users':
                            fk_columns.extend(fk.get('constrained_columns', []))
                    
                    for column in fk_columns:
                        if "_by" not in column:  # Skip columns already handled
                            # Try to nullify first
                            try:
                                query = text(f"UPDATE daily_feed_schedule SET {column} = NULL WHERE {column} = :user_id")
                                result = db.session.execute(query, {"user_id": user_id})
                                deletion_summary[f"daily_feed_schedule.{column}_nullified"] = result.rowcount
                                logger.info(f"Nullified {result.rowcount} records in daily_feed_schedule.{column}")
                                db.session.commit()
                            except Exception as e:
                                logger.warning(f"Could not nullify daily_feed_schedule.{column}: {str(e)}")
                                db.session.rollback()
                                
                                # If nullification fails, delete
                                query = text(f"DELETE FROM daily_feed_schedule WHERE {column} = :user_id")
                                result = db.session.execute(query, {"user_id": user_id})
                                deletion_summary[f"daily_feed_schedule.{column}_deleted"] = result.rowcount
                                logger.info(f"Deleted {result.rowcount} records from daily_feed_schedule where {column} = {user_id}")
                                db.session.commit()
            except Exception as e:
                logger.error(f"Error handling daily_feed_schedule records: {str(e)}")
                db.session.rollback()
                return jsonify({
                    "status": "error",
                    "message": "Failed to handle daily_feed_schedule records",
                    "details": str(e)
                }), 500
                
            # Process product_stock table and related notifications
            try:
                if "product_stock" in all_tables:
                    logger.info(f"Processing product_stock references to user {user_id}")
                    
                    # First get all product_stock IDs linked to this user's created_by_id
                    query = text("SELECT id FROM product_stock WHERE created_by_id = :user_id")
                    result = db.session.execute(query, {"user_id": user_id})
                    product_stock_ids = [row[0] for row in result.fetchall()]
                    
                    if product_stock_ids:
                        logger.info(f"Found {len(product_stock_ids)} product_stock records linked to user {user_id}")
                        
                        # Handle notifications that reference these product_stock records
                        if "notifications" in all_tables:
                            # Using IN clause with proper parameter binding for multiple values
                            placeholders = ', '.join([f':id{i}' for i in range(len(product_stock_ids))])
                            params = {f'id{i}': id_val for i, id_val in enumerate(product_stock_ids)}
                            
                            query = text(f"DELETE FROM notifications WHERE product_stock_id IN ({placeholders})")
                            result = db.session.execute(query, params)
                            deletion_summary["notifications.product_stock_id"] = result.rowcount
                            logger.info(f"Deleted {result.rowcount} notifications referring to user's product stocks")
                            db.session.commit()
                    
                    # Check for any not-yet-nullified columns
                    fk_columns = []
                    for fk in inspector.get_foreign_keys("product_stock"):
                        if fk.get('referred_table') == 'users':
                            fk_columns.extend(fk.get('constrained_columns', []))
                    
                    for column in fk_columns:
                        # Try to nullify first (even if already handled above, to be thorough)
                        try:
                            query = text(f"UPDATE product_stock SET {column} = NULL WHERE {column} = :user_id")
                            result = db.session.execute(query, {"user_id": user_id})
                            deletion_summary[f"product_stock.{column}_nullified"] = result.rowcount
                            logger.info(f"Nullified {result.rowcount} records in product_stock.{column}")
                            db.session.commit()
                        except Exception as e:
                            logger.warning(f"Could not nullify product_stock.{column}: {str(e)}")
                            db.session.rollback()
                            
                            # If nullification fails, delete
                            query = text(f"DELETE FROM product_stock WHERE {column} = :user_id")
                            result = db.session.execute(query, {"user_id": user_id})
                            deletion_summary[f"product_stock.{column}_deleted"] = result.rowcount
                            logger.info(f"Deleted {result.rowcount} records from product_stock where {column} = {user_id}")
                            db.session.commit()
            except Exception as e:
                logger.error(f"Error handling product_stock records: {str(e)}")
                db.session.rollback()
                return jsonify({
                    "status": "error", 
                    "message": "Failed to handle product stock records",
                    "details": str(e)
                }), 500
            
            # Process order_item records that reference product_type created by the user
            try:
                if "order_item" in all_tables and "product_type" in all_tables:
                    logger.info(f"Processing order_item references to product_type for user {user_id}")
                    
                    # Delete rows in order_item that reference product_type rows created by the user
                    query = text("""
                        DELETE FROM order_item 
                        WHERE product_type_id IN (
                            SELECT id FROM product_type WHERE created_by_id = :user_id
                        )
                    """)
                    result = db.session.execute(query, {"user_id": user_id})
                    deletion_summary["order_item.product_type_id_deleted"] = result.rowcount
                    logger.info(f"Deleted {result.rowcount} records from order_item referencing product_type created by user {user_id}")
                    db.session.commit()
            except Exception as e:
                logger.error(f"Error handling order_item records: {str(e)}")
                db.session.rollback()
                return jsonify({
                    "status": "error",
                    "message": "Failed to handle order_item records",
                    "details": str(e)
                }), 500
            
            # Process daily_feed_items
            try:
                if "daily_feed_items" in all_tables:
                    logger.info(f"Processing daily_feed_items references to user {user_id}")
                    
                    # Check for any not-yet-nullified columns
                    fk_columns = []
                    for fk in inspector.get_foreign_keys("daily_feed_items"):
                        if fk.get('referred_table') == 'users':
                            fk_columns.extend(fk.get('constrained_columns', []))
                    
                    for column in fk_columns:
                        if "_by" not in column:  # Skip columns already handled
                            # Try to nullify first
                            try:
                                query = text(f"UPDATE daily_feed_items SET {column} = NULL WHERE {column} = :user_id")
                                result = db.session.execute(query, {"user_id": user_id})
                                deletion_summary[f"daily_feed_items.{column}_nullified"] = result.rowcount
                                logger.info(f"Nullified {result.rowcount} records in daily_feed_items.{column}")
                                db.session.commit()
                            except Exception as e:
                                logger.warning(f"Could not nullify daily_feed_items.{column}: {str(e)}")
                                db.session.rollback()
                                
                                # If nullification fails, delete
                                query = text(f"DELETE FROM daily_feed_items WHERE {column} = :user_id")
                                result = db.session.execute(query, {"user_id": user_id})
                                deletion_summary[f"daily_feed_items.{column}_deleted"] = result.rowcount
                                logger.info(f"Deleted {result.rowcount} records from daily_feed_items where {column} = {user_id}")
                                db.session.commit()
            except Exception as e:
                logger.error(f"Error handling daily_feed_items records: {str(e)}")
                db.session.rollback()
                return jsonify({
                    "status": "error",
                    "message": "Failed to handle daily_feed_items records",
                    "details": str(e)
                }), 500
            
            # Process feed_stock table specifically to handle both updated_by and user_id
            try:
                if "feed_stock" in all_tables:
                    logger.info(f"Processing feed_stock references to user {user_id}")
                    
                    # Check for any not-yet-nullified columns
                    fk_columns = []
                    for fk in inspector.get_foreign_keys("feed_stock"):
                        if fk.get('referred_table') == 'users':
                            fk_columns.extend(fk.get('constrained_columns', []))
                    
                    # Process all columns that reference users.id
                    for column in fk_columns:
                        # Try to nullify first (even if already handled above, to be thorough)
                        try:
                            query = text(f"UPDATE feed_stock SET {column} = NULL WHERE {column} = :user_id")
                            result = db.session.execute(query, {"user_id": user_id})
                            deletion_summary[f"feed_stock.{column}_nullified"] = result.rowcount
                            logger.info(f"Nullified {result.rowcount} records in feed_stock.{column}")
                            db.session.commit()
                        except Exception as e:
                            logger.warning(f"Could not nullify feed_stock.{column}: {str(e)}")
                            db.session.rollback()
                            
                            # If nullification fails, delete those records
                            query = text(f"DELETE FROM feed_stock WHERE {column} = :user_id")
                            result = db.session.execute(query, {"user_id": user_id})
                            deletion_summary[f"feed_stock.{column}_deleted"] = result.rowcount
                            logger.info(f"Deleted {result.rowcount} records from feed_stock where {column} = {user_id}")
                            db.session.commit()
            except Exception as e:
                logger.error(f"Error handling feed_stock records: {str(e)}")
                db.session.rollback()
                return jsonify({
                    "status": "error", 
                    "message": "Failed to handle feed stock records",
                    "details": str(e)
                }), 500
            
            # Process all remaining tables
            logger.info("Step 3: Processing all remaining tables with foreign key constraints")
            for table_name in all_tables:
                # Skip special-handled tables
                if table_name in special_handling_tables:
                    continue
                
                try:
                    # Find all columns that might reference users.id
                    fk_columns = []
                    for fk in inspector.get_foreign_keys(table_name):
                        if fk.get('referred_table') == 'users':
                            fk_columns.extend(fk.get('constrained_columns', []))
                    
                    # For each column, delete remaining records (those that couldn't be nullified)
                    for column in fk_columns:
                        if "_by" not in column:  # Skip columns already handled in Step 1
                            logger.info(f"Processing {table_name}.{column} references to user {user_id}")
                            query = text(f"DELETE FROM {table_name} WHERE {column} = :user_id")
                            result = db.session.execute(query, {"user_id": user_id})
                            deletion_summary[f"{table_name}.{column}_deleted"] = result.rowcount
                            logger.info(f"Deleted {result.rowcount} records from {table_name} where {column} = {user_id}")
                            db.session.commit()
                
                except Exception as table_e:
                    logger.info(f"Skipped {table_name}: {str(table_e)}")
            
            # Finally delete the user
            logger.info(f"Step 4: Deleting user {user_id}")
            db.session.delete(user)
            db.session.commit()
            
            logger.info(f"User {user_id} successfully deleted with all relationships")
            return jsonify({
                "status": "success",
                "message": "User and all related data deleted successfully",
                "details": deletion_summary
            }), 200
            
        except Exception as inner_e:
            logger.error(f"Inner exception during deletion: {str(inner_e)}")
            logger.error(traceback.format_exc())
            db.session.rollback()
            error_msg = f"Inner deletion error: {str(inner_e)}"
            return jsonify({
                "status": "error", 
                "message": error_msg,
                "exception_type": type(inner_e).__name__,
                "details": str(inner_e)
            }), 500

    except Exception as e:
        logger.error(f"Outer exception during deletion: {str(e)}")
        logger.error(traceback.format_exc())
        if 'db' in locals():
            db.session.rollback()
        error_msg = f"Failed to delete user: {str(e)}"
        return jsonify({
            "status": "error", 
            "message": error_msg,
            "exception_type": type(e).__name__,
            "details": str(e)
        }), 500
@user_bp.route('/edit/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    try:
        # Cari user berdasarkan user_id
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Update user data
        data = request.json
        user.name = data.get("name", user.name)
        user.username = data.get("username", user.username)
        user.email = data.get("email", user.email)
        user.contact = data.get("contact", user.contact)
        user.religion = data.get("religion", user.religion)
        user.birth = data.get("birth", user.birth)
        user.role_id = data.get("role_id", user.role_id)

        db.session.commit()

        return jsonify({"message": "User updated successfully"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
    


@user_bp.route('/farmers', methods=['GET'])
def get_all_farmers():
    try:
        # Get all users with farmer role by joining User and Role tables
        farmers = User.query.join(Role, User.role_id == Role.id).filter(
            Role.name == 'farmer'  # Assuming 'farmer' is the role name for farmers
        ).add_columns(
            User.id,
            User.name,
            User.username,
            User.email,
            User.contact,
            User.religion,
            User.birth,
            Role.name.label('role_name')
        ).all()

        # Format farmers data into list of dictionaries
        farmers_list = [{
            "id": farmer.id,
            "name": farmer.name,
            "username": farmer.username,
            "email": farmer.email,
            "contact": farmer.contact,
            "religion": farmer.religion,
            "birth": farmer.birth,
            "role": farmer.role_name
        } for farmer in farmers]

        return jsonify({
            "status": "success",
            "message": "Farmers retrieved successfully",
            "total_farmers": len(farmers_list),
            "farmers": farmers_list
        }), 200

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    

@user_bp.route('/reset-password/<int:user_id>', methods=['POST'])
def reset_password(user_id):
    try:
        # Find user by ID
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404
        
        # Get the user's role
        role = Role.query.get(user.role_id)
        if not role:
            return jsonify({"error": "User role not found"}), 404
        
        # Set default password based on role
        default_password = None
        if role.name.lower() == 'farmer':
            default_password = 'farmer123'
        elif role.name.lower() == 'supervisor':
            default_password = 'supervisor123'
        else:
            return jsonify({"error": f"Password reset not supported for role: {role.name}"}), 400
        
        # Hash the default password
        hashed_password = generate_password_hash(default_password)
        
        # Update user password
        user.password = hashed_password
        db.session.commit()
        
        return jsonify({
            "status": "success",
            "message": f"Password reset successfully for {user.name}",
            "user_id": user.id,
            "role": role.name
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    
@user_bp.route('/change-password/<int:user_id>', methods=['POST'])
def change_password(user_id):
    try:
        data = request.get_json()
        old_password = data.get('old_password')
        new_password = data.get('new_password')

        if not old_password or not new_password:
            return jsonify({"status": "error", "message": "Old and new password are required"}), 400

        user = User.query.get(user_id)
        if not user:
            return jsonify({"status": "error", "message": "User not found"}), 404

        # Verifikasi password lama
        if not check_password_hash(user.password, old_password):
            return jsonify({"status": "error", "message": "Old password is incorrect"}), 400

        # Update password baru
        user.password = generate_password_hash(new_password)
        db.session.commit()

        return jsonify({"status": "success", "message": "Password changed successfully"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
